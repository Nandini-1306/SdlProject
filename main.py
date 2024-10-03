

from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import os
import csv
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)
CORS(app)

UPLOAD_FOLDER = 'uploads'
RESULTS_BASE_FOLDER = 'results'
CSV_BASE_FOLDER = 'csv'
EXCEL_BASE_FOLDER = 'excel'

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def create_folders_for_pdf(pdf_name):
    base_name = os.path.splitext(pdf_name)[0]
    results_folder = os.path.join(RESULTS_BASE_FOLDER, base_name)
    csv_folder = os.path.join(CSV_BASE_FOLDER, base_name)
    excel_folder = os.path.join(EXCEL_BASE_FOLDER, base_name)
    
    os.makedirs(results_folder, exist_ok=True)
    os.makedirs(csv_folder, exist_ok=True)
    os.makedirs(excel_folder, exist_ok=True)
    
    return results_folder, csv_folder, excel_folder

def extract_data_to_csv(pdf_path, csv_folder):
    csv_path = os.path.join(csv_folder, 'data.csv')
    with pdfplumber.open(pdf_path) as pdf:
        with open(csv_path, 'w', newline='') as csv_file:
            writer = csv.writer(csv_file)
            for page in pdf.pages:
                table = page.extract_table()
                if table:
                    writer.writerows(table)
    return csv_path

def highlight_attendance(csv_path, excel_folder, subject, attendance_type):
    # Load the CSV file
    df = pd.read_csv(csv_path, skiprows=5)
    
    # Log the DataFrame columns for debugging
    print("DataFrame Columns:", df.columns.tolist())

    # Find the column indices for the specified subject
    subject_columns = df.columns.tolist()

    # Log the input subject for debugging
    print('Input Subject:', subject)  # Log the received subject
    df.columns = df.columns.str.strip()

    # Strip whitespace from the subject input
    subject = subject.strip().upper()
    
    # Initialize the column variables
    total_classes_col = None
    present_classes_col = None
    percentage_col = None

    # Determine column positions based on attendance type
    for i in range(len(subject_columns)):
        if subject_columns[i].strip().upper() == subject:
            if attendance_type == 'TH':
                total_classes_col = subject_columns[i]
                present_classes_col = subject_columns[i + 1]
                percentage_col = subject_columns[i + 2]
            elif attendance_type == 'LAB':
                if i + 3 < len(subject_columns) and subject_columns[i + 3].startswith("Unnamed"):
                    total_classes_col = subject_columns[i + 3]
                    present_classes_col = subject_columns[i + 4]
                    percentage_col = subject_columns[i + 5]
                else:
                    raise ValueError(f"The subject {subject} does not have Lab attendance data.")
            break

    # Log the found column names for debugging
    print('Columns:', total_classes_col, present_classes_col, percentage_col)

    # Highlight cells with attendance < 60%
    if total_classes_col and present_classes_col and percentage_col:
        # Write DataFrame to an Excel file
        output_excel = os.path.join(excel_folder, f'{subject}_highlighted_attendance.xlsx')
        df.to_excel(output_excel, index=False)

        # Load the Excel file with openpyxl
        wb = load_workbook(output_excel)
        ws = wb.active

        # Yellow fill for cells with less than 60% attendance
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Highlight cells with less than 60% attendance in the percentage column
        for row in range(2, ws.max_row + 1):  # Start from row 2 to skip the header
            percentage_value = ws.cell(row=row, column=subject_columns.index(percentage_col) + 1).value
            if percentage_value is not None and isinstance(percentage_value, (int, float)):
                if percentage_value < 60:
                    ws.cell(row=row, column=subject_columns.index(percentage_col) + 1).fill = yellow_fill

        # Save the workbook
        wb.save(output_excel)
        return output_excel
    else:
        raise ValueError(f"Attendance data for subject {subject} not found.")



@app.route('/upload', methods=['POST'])
def upload_file():
    if 'pdfFile' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    
    file = request.files['pdfFile']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    if file and file.filename.endswith('.pdf'):
        file_path = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(file_path)
        
        try:
            # Process the PDF and convert it to CSV
            results_folder, csv_folder, excel_folder = create_folders_for_pdf(file.filename)
            csv_path = extract_data_to_csv(file_path, csv_folder)
            
            # Get user inputs for subject and attendance type
            subject = request.form.get('subject').strip().upper()
            attendance_type = request.form.get('attendance_type').strip().upper()

            # Log the subject and attendance type for debugging
            print('Subject:', subject)  # Log the received subject
            print('Attendance Type:', attendance_type)  # Log the received attendance type

            # Highlight attendance data
            highlighted_excel = highlight_attendance(csv_path, excel_folder, subject, attendance_type)

            return jsonify({'csv_file': 'data.csv', 'highlighted_excel': highlighted_excel}), 200
        except ValueError as e:
            return jsonify({'error': str(e)}), 400
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    return jsonify({'error': 'Invalid file format'}), 400


@app.route('/download/<filename>')
def download_file(filename):
    safe_filename = os.path.basename(filename)  # Prevent directory traversal
    base_name = os.path.splitext(filename)[0]
    results_folder = os.path.join(RESULTS_BASE_FOLDER, base_name)
    if filename.endswith('.pdf'):
        return send_from_directory(results_folder, safe_filename)
    elif filename.endswith('.xlsx'):
        excel_folder = os.path.join(EXCEL_BASE_FOLDER, base_name)
        return send_from_directory(excel_folder, safe_filename)
    else:
        return jsonify({'error': 'File not found'}), 404

if __name__ == '__main__':
    app.run(debug=True)
