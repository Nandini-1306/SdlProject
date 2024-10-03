import os
import csv
import pdfplumber
import pytesseract
import pandas as pd
from PIL import Image
from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from werkzeug.utils import secure_filename

app = Flask(__name__)
CORS(app)

UPLOAD_FOLDER = 'uploads'
RESULTS_BASE_FOLDER = 'results'
CSV_BASE_FOLDER = 'csv'
EXCEL_BASE_FOLDER = 'excel'

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def create_folders_for_file(file_name):
    base_name = os.path.splitext(file_name)[0]
    results_folder = os.path.join(RESULTS_BASE_FOLDER, base_name)
    csv_folder = os.path.join(CSV_BASE_FOLDER, base_name)
    excel_folder = os.path.join(EXCEL_BASE_FOLDER, base_name)
    
    os.makedirs(results_folder, exist_ok=True)
    os.makedirs(csv_folder, exist_ok=True)
    os.makedirs(excel_folder, exist_ok=True)
    
    return results_folder, csv_folder, excel_folder

def extract_data_from_pdf(pdf_path, csv_folder):
    csv_path = os.path.join(csv_folder, 'data.csv')
    with pdfplumber.open(pdf_path) as pdf:
        with open(csv_path, 'w', newline='') as csv_file:
            writer = csv.writer(csv_file)
            for page in pdf.pages:
                table = page.extract_table()
                if table:
                    writer.writerows(table)
    print(f"Extracted data from PDF to {csv_path}")  # Log the CSV path
    return csv_path

def extract_data_from_image_pdf(pdf_path, csv_folder):
    from pdf2image import convert_from_path

    # Convert PDF pages to images
    images = convert_from_path(pdf_path)
    extracted_data = []

    for i, image in enumerate(images):
        image_path = os.path.join(csv_folder, f'page_{i + 1}.png')
        image.save(image_path, 'PNG')

        # Perform OCR on the image
        text = pytesseract.image_to_string(image)

        # Optional: Process the text here to extract structured data
        # For example, you can split the text into lines and process accordingly
        lines = text.splitlines()
        for line in lines:
            if line.strip():  # Skip empty lines
                extracted_data.append(line.split())  # Adjust this to fit your desired CSV format

    # Save OCR results into a CSV file
    csv_path = os.path.join(csv_folder, 'ocr_data.csv')
    with open(csv_path, 'w', newline='') as csv_file:
        writer = csv.writer(csv_file)
        writer.writerows(extracted_data)  # Write the structured data

    print(f"Extracted OCR data to {csv_path}")  # Log the CSV path
    return csv_path

def highlight_attendance(csv_path, excel_folder, subject, attendance_type, highlight_last_column=False):
    # Load the CSV file

    
    df = pd.read_csv(csv_path, skiprows=5)

    # Log the DataFrame for debugging
    print("DataFrame Columns:", df.columns.tolist())
    print("DataFrame Head:\n", df.head())  # Log the first few rows of the DataFrame

    # Clean subject input
    subject = subject.strip().upper()

    # Check if the subject is "TOTAL" and set flag
    if subject == "TOTAL":
        highlight_last_column = True  # Set this flag to True to highlight the last column

    # Write DataFrame to Excel
    output_excel = os.path.join(excel_folder, f'{subject}_highlighted_attendance.xlsx')
    df.to_excel(output_excel, index=False)
    print(f"Written DataFrame to Excel: {output_excel}")  # Log the Excel path

    # Load the Excel file with openpyxl for styling
    wb = load_workbook(output_excel)
    ws = wb.active

    # Yellow fill for cells with less than 60% attendance
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # If highlighting the last column
    if highlight_last_column:
        percentage_col = df.columns[-1]  # Last column for highlighting
        percentage_idx = df.columns.get_loc(percentage_col) + 1  # Get 1-based index
        
        # Highlight cells with attendance < 60% in the last column
        for row in range(2, ws.max_row + 1):  # Skip the header
            percentage_value = ws.cell(row=row, column=percentage_idx).value
            if isinstance(percentage_value, (int, float)) and percentage_value < 60:
                ws.cell(row=row, column=percentage_idx).fill = yellow_fill
                print(f"Highlighted row {row} in column {percentage_col} with value {percentage_value}")  # Log highlighting
    else:
        # Find the specific attendance columns
        total_classes_col, present_classes_col, percentage_col = None, None, None
        subject_columns = df.columns.tolist()

        for i in range(len(subject_columns)):
            if subject_columns[i].strip().upper() == subject:
                if attendance_type == 'TH':
                    total_classes_col = subject_columns[i]
                    present_classes_col = subject_columns[i + 1]
                    percentage_col = subject_columns[i + 2]
                elif attendance_type == 'LAB':
                    total_classes_col = subject_columns[i + 3]
                    present_classes_col = subject_columns[i + 4]
                    percentage_col = subject_columns[i + 5]
                break

        if not total_classes_col or not present_classes_col or not percentage_col:
            raise ValueError(f"Attendance data for subject {subject} not found.")
        
        print(f"Columns found - Total Classes: {total_classes_col}, Present Classes: {present_classes_col}, Percentage: {percentage_col}")  # Log column info
        
        # Highlight cells with attendance < 60% in the specific percentage column
        percentage_idx = df.columns.get_loc(percentage_col) + 1  # Get 1-based index
        for row in range(2, ws.max_row + 1):  # Skip the header
            percentage_value = ws.cell(row=row, column=percentage_idx).value
            if isinstance(percentage_value, (int, float)) and percentage_value < 60:
                ws.cell(row=row, column=percentage_idx).fill = yellow_fill
                print(f"Highlighted row {row} in column {percentage_col} with value {percentage_value}")  # Log highlighting

    wb.save(output_excel)  # Ensure the workbook is saved
    print(f"Saved highlighted Excel file: {output_excel}")  # Log the final save
    return output_excel

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    filename = secure_filename(file.filename)
    file_path = os.path.join(UPLOAD_FOLDER, filename)
    file.save(file_path)
    print(f"File uploaded: {file_path}")  # Log the upload

    try:
        # Create necessary folders for processing
        results_folder, csv_folder, excel_folder = create_folders_for_file(filename)
        
        # Process based on file type
        if filename.endswith('.pdf'):
            # Use pdfplumber to check if it's a text-based or image-based PDF
            try:
                csv_path = extract_data_from_pdf(file_path, csv_folder)
            except Exception:
                csv_path = extract_data_from_image_pdf(file_path, csv_folder)
        
        elif filename.endswith('.xlsx') or filename.endswith('.xls'):
            # Directly handle Excel files
            df = pd.read_excel(file_path)
            csv_path = os.path.join(csv_folder, 'data.csv')
            df.to_csv(csv_path, index=False)
            print(f"Converted Excel to CSV: {csv_path}")  # Log the conversion

        else:
            return jsonify({'error': 'Unsupported file format'}), 400

        # Get user input for subject and attendance type
        subject = request.form.get('subject')
        attendance_type = request.form.get('attendance_type')
        highlight_last_column = 'highlight_last_column' in request.form and request.form['highlight_last_column'] == 'true'

        # Clean subject input
        subject = subject.strip().upper()

        # Handle the case when "Total" is the subject
        if subject == "TOTAL":
            attendance_type = None  # Set attendance_type to None if subject is "Total"
        else:
            if attendance_type is None:
                return jsonify({'error': 'Attendance type is required when subject is not "Total".'}), 400
            
            attendance_type = attendance_type.strip().upper()

        # Log the subject, attendance type, and CSV path
        print(f"Subject: {subject}")
        print(f"Attendance Type: {attendance_type}")
        print(f"CSV Path: {csv_path}")

        # Highlight attendance data
        highlighted_excel = highlight_attendance(csv_path, excel_folder, subject, attendance_type, highlight_last_column)

        return jsonify({'csv_file': 'data.csv', 'highlighted_excel': highlighted_excel}), 200

    except ValueError as e:
        print(f"ValueError: {str(e)}")  # Log ValueErrorx
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        print(f"Exception: {str(e)}")  # Log generic Exception
        return jsonify({'error': str(e)}), 500


@app.route('/download/<filename>')
def download_file(filename):
    safe_filename = os.path.basename(filename)
    base_name = os.path.splitext(filename)[0]
    if filename.endswith('.pdf'):
        results_folder = os.path.join(RESULTS_BASE_FOLDER, base_name)
        return send_from_directory(results_folder, safe_filename)
    elif filename.endswith('.xlsx'):
        excel_folder = os.path.join(EXCEL_BASE_FOLDER, base_name)
        return send_from_directory(excel_folder, safe_filename)
    return jsonify({'error': 'File not found'}), 404

if __name__ == '__main__':
    app.run(debug=True)
