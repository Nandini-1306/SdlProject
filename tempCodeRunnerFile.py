import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the CSV file
file_path = r'C:\Users\nandi\Downloads\data (2).csv'
df = pd.read_csv(file_path, skiprows=5)

# User inputs
subject = input("Enter the subject (e.g., DBMS, CN, TOC, etc.): ").strip().upper()
attendance_type = input("Enter the attendance type (TH for Theory, LAB for Lab, TOTAL): ").strip().upper()

# Find the column indices for the specified subject
subject_columns = df.columns.tolist()

# Initialize the column variables
total_classes_col = None
present_classes_col = None
percentage_col = None

# Iterate over the columns to find the correct positions
for i in range(len(subject_columns)):
    if subject_columns[i].strip().upper() == subject:
        if attendance_type == 'TH':
            total_classes_col = subject_columns[i]
            present_classes_col = subject_columns[i + 1]
            percentage_col = subject_columns[i + 2]
        elif attendance_type == 'LAB':
            if i + 3 < len(subject_columns) and subject_columns[i + 3].startswith("Unnamed"):
                total_classes_col = subject_columns[i + 3]
                present_classes_col = subject_columns[i + 4]
                percentage_col = subject_columns[i + 5]
            else:
                print(f"Error: The subject {subject} does not have Lab attendance data.")
                break
        break

# Show attendance data and highlight cells with <60% attendance
if total_classes_col and present_classes_col and percentage_col:
    try:
        print(f"\nAttendance Data for {subject} - {attendance_type.capitalize()} Attendance:")
        
        # Write DataFrame to an Excel file
        output_excel = r'C:\Users\nandi\Downloads\highlighted_attendance.xlsx'
        df.to_excel(output_excel, index=False)

        # Load the Excel file with openpyxl
        wb = load_workbook(output_excel)
        ws = wb.active

        # Yellow fill for cells with less than 60% attendance
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Find and highlight cells with less than 60% attendance in the percentage column
        for row in range(2, ws.max_row + 1):  # Start from row 2 to skip the header
            percentage_value = ws.cell(row=row, column=subject_columns.index(percentage_col) + 1).value
            if percentage_value is not None and isinstance(percentage_value, (int, float)):
                if percentage_value < 60:
                    ws.cell(row=row, column=subject_columns.index(percentage_col) + 1).fill = yellow_fill

        # Save the workbook
        wb.save(output_excel)
        print(f"Highlighted cells saved to {output_excel}")

    except KeyError as e:
        print(f"Error: {e}. One or more columns not found in the CSV.")
else:
    print(f"Attendance data for subject {subject} not found.")
